"""Task API — §8.4 (Branch 07–09: list, create, update, init-week, export, distribution)."""
import csv
import io
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, File, Query, UploadFile
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.responses import JSONResponse, StreamingResponse

from app.api.deps import get_current_user, get_db
from app.models.reference import Aircraft, WorkPackage
from app.models.shop import Shop
from app.models.task import TaskItem, TaskSnapshot
from app.models.user import Role, User, user_roles
from app.models.user_shop_access import UserShopAccess
from app.schemas.common import APIError, PaginatedResponse, pagination_params
from app.schemas.task import (
    AssignRequest,
    AssignWorkerRequest,
    BatchUpdateRequest,
    BatchUpdateResponse,
    BulkAssignRequest,
    ImportConfirmRequest,
    InitWeekRequest,
    InitWeekResponse,
    SnapshotDeleteResponse,
    SnapshotListItem,
    SnapshotRestoreResponse,
    SnapshotUpdate,
    SnapshotUpdateResponse,
    SnapshotVersionRequest,
    TaskCreate,
    TaskCreateResponse,
    TaskDeactivateResponse,
)
from app.services.audit_service import write_audit
from app.services.shop_access_service import enforce_shop_access
from app.services.task_service import (
    check_mh_decrease,
    validate_status,
    init_week,
)

router = APIRouter(prefix="/api/tasks", tags=["tasks"])


# ── Helpers ───────────────────────────────────────────────────────────

def _user_name(users_map: dict, uid: int | None) -> str | None:
    if uid is None:
        return None
    u = users_map.get(uid)
    return u.name if u else None


async def _build_users_map(db: AsyncSession, ids: set[int]) -> dict[int, User]:
    if not ids:
        return {}
    rows = (await db.execute(select(User).where(User.id.in_(ids)))).scalars().all()
    return {u.id: u for u in rows}


def _snap_to_dict(snap: TaskSnapshot) -> dict:
    return {
        "snapshot_id": snap.id,
        "version": snap.version,
        "status": snap.status,
        "mh_incurred_hours": str(snap.mh_incurred_hours),
        "deadline_date": str(snap.deadline_date) if snap.deadline_date else None,
        "remarks": snap.remarks,
        "critical_issue": snap.critical_issue,
        "has_issue": snap.has_issue,
        "correction_reason": snap.correction_reason,
        "last_updated_at": str(snap.last_updated_at),
        "last_updated_by": snap.last_updated_by,
        "supervisor_updated_at": str(snap.supervisor_updated_at) if snap.supervisor_updated_at else None,
        "is_deleted": snap.is_deleted,
        "meeting_date": str(snap.meeting_date),
        "task_id": snap.task_id,
    }


def _require_admin(current_user: dict) -> None:
    """Raise 403 if current_user is not ADMIN."""
    if "ADMIN" not in current_user.get("roles", []):
        raise APIError(403, "Insufficient permissions", "FORBIDDEN")


def _normalize_airline_category(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip().upper()
    if normalized == "3RD":
        normalized = "THIRD_PARTIES"
    if normalized not in {"ALL", "SQ", "THIRD_PARTIES"}:
        raise APIError(
            422,
            "airline_category must be one of ALL, SQ, THIRD_PARTIES",
            "VALIDATION_ERROR",
            field="airline_category",
        )
    return normalized


async def _validate_assigned_supervisor(
    db: AsyncSession,
    *,
    assigned_supervisor_id: int,
    shop_id: int,
) -> None:
    """Ensure assignment target is an active SUPERVISOR with access to shop."""
    supervisor_id = (
        await db.execute(
            select(User.id)
            .join(user_roles, User.id == user_roles.c.user_id)
            .join(Role, Role.id == user_roles.c.role_id)
            .where(
                User.id == assigned_supervisor_id,
                User.is_active == True,
                Role.name == "SUPERVISOR",
            )
        )
    ).scalar_one_or_none()
    if supervisor_id is None:
        raise APIError(
            422,
            "assigned_supervisor_id must reference an active SUPERVISOR",
            "VALIDATION_ERROR",
            field="assigned_supervisor_id",
        )

    has_access = (
        await db.execute(
            select(UserShopAccess.id).where(
                UserShopAccess.user_id == assigned_supervisor_id,
                UserShopAccess.shop_id == shop_id,
            )
        )
    ).scalar_one_or_none()
    if has_access is None:
        raise APIError(
            403,
            "Supervisor does not have access to this shop",
            "SHOP_ACCESS_DENIED",
        )


# ── 1. GET /api/tasks/snapshots ──────────────────────────────────────

@router.get("/snapshots", response_model=PaginatedResponse[SnapshotListItem])
async def list_snapshots(
    meeting_date: date = Query(...),
    shop_id: int = Query(...),
    include_deleted: bool = Query(False),
    work_package_id: int | None = Query(None),
    assigned_supervisor_id: int | None = Query(None),
    aircraft_id: int | None = Query(None),
    status: str | None = Query(None),
    has_issue: bool | None = Query(None),
    airline_category: str | None = Query(None),
    paging: dict = Depends(pagination_params),
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await enforce_shop_access(db, current_user, shop_id, "VIEW")
    normalized_airline_category = _normalize_airline_category(airline_category)

    if status is not None:
        status = validate_status(status)

    # Base query: join snapshots with task_items
    q = (
        select(TaskSnapshot, TaskItem)
        .join(TaskItem, TaskSnapshot.task_id == TaskItem.id)
        .where(
            TaskSnapshot.meeting_date == meeting_date,
            TaskItem.shop_id == shop_id,
        )
    )
    cq = (
        select(func.count())
        .select_from(TaskSnapshot)
        .join(TaskItem, TaskSnapshot.task_id == TaskItem.id)
        .where(
            TaskSnapshot.meeting_date == meeting_date,
            TaskItem.shop_id == shop_id,
        )
    )

    if not include_deleted:
        q = q.where(TaskSnapshot.is_deleted == False)
        cq = cq.where(TaskSnapshot.is_deleted == False)

    if work_package_id is not None:
        q = q.where(TaskItem.work_package_id == work_package_id)
        cq = cq.where(TaskItem.work_package_id == work_package_id)

    if assigned_supervisor_id is not None:
        q = q.where(TaskItem.assigned_supervisor_id == assigned_supervisor_id)
        cq = cq.where(TaskItem.assigned_supervisor_id == assigned_supervisor_id)

    if aircraft_id is not None:
        q = q.where(TaskItem.aircraft_id == aircraft_id)
        cq = cq.where(TaskItem.aircraft_id == aircraft_id)

    if status is not None:
        q = q.where(TaskSnapshot.status == status)
        cq = cq.where(TaskSnapshot.status == status)

    if has_issue is not None:
        q = q.where(TaskSnapshot.has_issue == has_issue)
        cq = cq.where(TaskSnapshot.has_issue == has_issue)

    if normalized_airline_category and normalized_airline_category != "ALL":
        sq_condition = func.lower(func.trim(func.coalesce(Aircraft.airline, ""))).in_(
            ["sq", "singapore airlines"]
        )
        q = q.join(Aircraft, Aircraft.id == TaskItem.aircraft_id)
        cq = cq.join(Aircraft, Aircraft.id == TaskItem.aircraft_id)
        if normalized_airline_category == "SQ":
            q = q.where(sq_condition)
            cq = cq.where(sq_condition)
        else:
            q = q.where(~sq_condition)
            cq = cq.where(~sq_condition)

    total = (await db.execute(cq)).scalar() or 0

    rows = (
        await db.execute(
            q.order_by(TaskSnapshot.id)
            .offset(paging["offset"])
            .limit(paging["per_page"])
        )
    ).all()

    # Collect IDs for batch resolution
    user_ids: set[int] = set()
    ac_ids: set[int] = set()
    wp_ids: set[int] = set()
    shop_ids: set[int] = set()

    for snap, task in rows:
        user_ids.add(snap.last_updated_by)
        if task.assigned_supervisor_id:
            user_ids.add(task.assigned_supervisor_id)
        if task.assigned_worker_id:
            user_ids.add(task.assigned_worker_id)
        ac_ids.add(task.aircraft_id)
        if task.work_package_id:
            wp_ids.add(task.work_package_id)
        shop_ids.add(task.shop_id)

    users_map = await _build_users_map(db, user_ids)

    ac_map: dict[int, Aircraft] = {}
    if ac_ids:
        acs = (await db.execute(select(Aircraft).where(Aircraft.id.in_(ac_ids)))).scalars().all()
        ac_map = {a.id: a for a in acs}

    wp_map: dict[int, WorkPackage] = {}
    if wp_ids:
        wps = (await db.execute(select(WorkPackage).where(WorkPackage.id.in_(wp_ids)))).scalars().all()
        wp_map = {w.id: w for w in wps}

    shop_map: dict[int, Shop] = {}
    if shop_ids:
        shops = (await db.execute(select(Shop).where(Shop.id.in_(shop_ids)))).scalars().all()
        shop_map = {s.id: s for s in shops}

    items = []
    for snap, task in rows:
        ac = ac_map.get(task.aircraft_id)
        wp = wp_map.get(task.work_package_id) if task.work_package_id else None
        shop = shop_map.get(task.shop_id)

        items.append({
            "snapshot_id": snap.id,
            "task_id": task.id,
            "meeting_date": snap.meeting_date,
            "aircraft_id": task.aircraft_id,
            "work_package_id": task.work_package_id,
            "rfo_no": wp.rfo_no if wp else None,
            "ac_reg": ac.ac_reg if ac else "",
            "shop_id": task.shop_id,
            "shop_name": shop.name if shop else "",
            "assigned_supervisor_id": task.assigned_supervisor_id,
            "assigned_supervisor_name": _user_name(users_map, task.assigned_supervisor_id),
            "assigned_worker_id": task.assigned_worker_id,
            "assigned_worker_name": _user_name(users_map, task.assigned_worker_id),
            "distributed_at": task.distributed_at,
            "planned_mh": task.planned_mh,
            "task_text": task.task_text,
            "status": snap.status,
            "mh_incurred_hours": snap.mh_incurred_hours,
            "remarks": snap.remarks,
            "critical_issue": snap.critical_issue,
            "has_issue": snap.has_issue,
            "deadline_date": snap.deadline_date,
            "correction_reason": snap.correction_reason,
            "is_deleted": snap.is_deleted,
            "version": snap.version,
            "supervisor_updated_at": snap.supervisor_updated_at,
            "last_updated_at": snap.last_updated_at,
            "last_updated_by": snap.last_updated_by,
            "is_active": task.is_active,
        })

    return {
        "items": items,
        "total": total,
        "page": paging["page"],
        "per_page": paging["per_page"],
    }


# ── 2. GET /api/tasks/export/csv ─────────────────────────────────────
# MUST be registered BEFORE any /{task_id} path parameter routes.

@router.get("/export/csv")
async def export_csv(
    meeting_date: date = Query(...),
    shop_id: int = Query(...),
    include_deleted: bool = Query(False),
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await enforce_shop_access(db, current_user, shop_id, "VIEW")

    # Main query: join snapshots + task_items
    q = (
        select(TaskSnapshot, TaskItem)
        .join(TaskItem, TaskSnapshot.task_id == TaskItem.id)
        .where(
            TaskSnapshot.meeting_date == meeting_date,
            TaskItem.shop_id == shop_id,
        )
    )

    if not include_deleted:
        q = q.where(TaskSnapshot.is_deleted == False)

    rows = (await db.execute(q.order_by(TaskSnapshot.id))).all()

    # Collect IDs for batch lookup
    user_ids: set[int] = set()
    wp_ids: set[int] = set()
    task_ids_for_prev: set[int] = set()

    for snap, task in rows:
        user_ids.add(snap.last_updated_by)
        if task.assigned_supervisor_id:
            user_ids.add(task.assigned_supervisor_id)
        if task.assigned_worker_id:
            user_ids.add(task.assigned_worker_id)
        if task.work_package_id:
            wp_ids.add(task.work_package_id)
        task_ids_for_prev.add(snap.task_id)

    users_map = await _build_users_map(db, user_ids)

    wp_map: dict[int, WorkPackage] = {}
    if wp_ids:
        wps = (await db.execute(select(WorkPackage).where(WorkPackage.id.in_(wp_ids)))).scalars().all()
        wp_map = {w.id: w for w in wps}

    # Previous week snapshots for weekly_mh_delta calculation
    prev_date = meeting_date - timedelta(days=7)
    prev_mh_map: dict[int, Decimal] = {}
    if task_ids_for_prev:
        prev_snaps = (
            await db.execute(
                select(TaskSnapshot.task_id, TaskSnapshot.mh_incurred_hours)
                .where(
                    TaskSnapshot.task_id.in_(task_ids_for_prev),
                    TaskSnapshot.meeting_date == prev_date,
                )
            )
        ).all()
        for task_id_val, mh_val in prev_snaps:
            prev_mh_map[task_id_val] = mh_val

    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output)

    columns = [
        "task_id", "task_text", "status", "mh_incurred_hours",
        "work_package_id", "rfo_no", "assigned_supervisor_name",
        "assigned_worker_name", "distributed_at", "planned_mh",
        "weekly_mh_delta", "deadline_date", "remarks", "critical_issue",
        "has_issue", "updated_by_name", "updated_at",
    ]
    writer.writerow(columns)

    for snap, task in rows:
        wp = wp_map.get(task.work_package_id) if task.work_package_id else None
        current_mh = snap.mh_incurred_hours or Decimal("0")
        prev_mh = prev_mh_map.get(snap.task_id)
        weekly_mh_delta = current_mh - prev_mh if prev_mh is not None else current_mh

        writer.writerow([
            task.id,
            task.task_text,
            snap.status,
            str(snap.mh_incurred_hours),
            task.work_package_id or "",
            wp.rfo_no if wp else "",
            _user_name(users_map, task.assigned_supervisor_id) or "",
            _user_name(users_map, task.assigned_worker_id) or "",
            str(task.distributed_at) if task.distributed_at else "",
            str(task.planned_mh) if task.planned_mh else "",
            str(weekly_mh_delta),
            str(snap.deadline_date) if snap.deadline_date else "",
            snap.remarks or "",
            snap.critical_issue or "",
            snap.has_issue,
            _user_name(users_map, snap.last_updated_by) or "",
            str(snap.last_updated_at),
        ])

    output.seek(0)
    filename = f"tasks_{meeting_date}_{shop_id}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── 3. PATCH /api/tasks/snapshots/batch ──────────────────────────────
# NOTE: Must be registered before /snapshots/{snapshot_id} to avoid
#       FastAPI matching "batch" as a path parameter.

@router.patch("/snapshots/batch", response_model=BatchUpdateResponse)
async def batch_update_snapshots(
    body: BatchUpdateRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if not body.updates:
        raise APIError(422, "updates list must not be empty", "VALIDATION_ERROR")

    errors: list[dict] = []
    results: list[dict] = []

    # Pre-load all snapshots + tasks in one pass
    snap_ids = [u.snapshot_id for u in body.updates]
    snap_rows = (
        await db.execute(select(TaskSnapshot).where(TaskSnapshot.id.in_(snap_ids)))
    ).scalars().all()
    snap_map = {s.id: s for s in snap_rows}

    task_ids = {s.task_id for s in snap_rows}
    task_rows = (
        await db.execute(select(TaskItem).where(TaskItem.id.in_(task_ids)))
    ).scalars().all()
    task_map = {t.id: t for t in task_rows}

    # Check shop access for all involved shops
    shop_ids_involved = {t.shop_id for t in task_rows}
    for sid in shop_ids_involved:
        await enforce_shop_access(db, current_user, sid, "EDIT")

    now = datetime.now(timezone.utc)

    for item in body.updates:
        snap = snap_map.get(item.snapshot_id)
        if not snap:
            errors.append({
                "snapshot_id": item.snapshot_id,
                "code": "NOT_FOUND",
                "detail": "Snapshot not found",
            })
            continue

        task = task_map.get(snap.task_id)
        if not task:
            errors.append({
                "snapshot_id": item.snapshot_id,
                "code": "NOT_FOUND",
                "detail": "Task not found",
            })
            continue

        # Version check
        if snap.version != item.version:
            errors.append({
                "snapshot_id": item.snapshot_id,
                "code": "CONFLICT_VERSION",
                "current_version": snap.version,
            })
            continue

        # Validate status
        if item.status is not None:
            try:
                validate_status(item.status)
            except APIError:
                errors.append({
                    "snapshot_id": item.snapshot_id,
                    "code": "VALIDATION_ERROR",
                    "detail": f"Invalid status: {item.status}",
                })
                continue

        # MH decrease check
        if item.mh_incurred_hours is not None:
            try:
                await check_mh_decrease(
                    db, snap, item.mh_incurred_hours, current_user,
                    task.shop_id, item.correction_reason,
                )
            except APIError as e:
                errors.append({
                    "snapshot_id": item.snapshot_id,
                    "code": e.code,
                    "detail": e.detail,
                })
                continue

    # All-or-nothing: if any errors, rollback entire batch
    if errors:
        await db.rollback()
        return JSONResponse(
            status_code=422,
            content={
                "detail": "Batch update failed. All changes rolled back.",
                "code": "BATCH_VALIDATION_ERROR",
                "errors": errors,
            },
        )

    # Apply changes
    for item in body.updates:
        snap = snap_map[item.snapshot_id]
        before = _snap_to_dict(snap)

        if item.status is not None:
            snap.status = validate_status(item.status)
        if item.mh_incurred_hours is not None:
            snap.mh_incurred_hours = item.mh_incurred_hours
        if item.has_issue is not None:
            snap.has_issue = item.has_issue
        if item.remarks is not None:
            snap.remarks = item.remarks
        if item.critical_issue is not None:
            snap.critical_issue = item.critical_issue
        if item.correction_reason is not None:
            snap.correction_reason = item.correction_reason
        if "deadline_date" in item.model_fields_set:
            snap.deadline_date = item.deadline_date

        snap.version += 1
        snap.last_updated_at = now
        snap.last_updated_by = current_user["user_id"]
        snap.supervisor_updated_at = now

        await db.flush()

        await write_audit(
            db,
            actor_id=current_user["user_id"],
            entity_type="task_snapshot",
            entity_id=snap.id,
            action="UPDATE",
            before=before,
            after=_snap_to_dict(snap),
        )

        results.append({
            "snapshot_id": snap.id,
            "version": snap.version,
            "status": snap.status,
            "mh_incurred_hours": snap.mh_incurred_hours,
            "deadline_date": snap.deadline_date,
            "remarks": snap.remarks,
            "critical_issue": snap.critical_issue,
            "has_issue": snap.has_issue,
            "correction_reason": snap.correction_reason,
            "last_updated_at": snap.last_updated_at,
            "last_updated_by": snap.last_updated_by,
            "supervisor_updated_at": snap.supervisor_updated_at,
        })

    await db.commit()
    return {"items": results}


# ── 4. PATCH /api/tasks/snapshots/{snapshot_id} ─────────────────────

@router.patch("/snapshots/{snapshot_id}", response_model=SnapshotUpdateResponse)
async def update_snapshot(
    snapshot_id: int,
    body: SnapshotUpdate,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    snap = (
        await db.execute(select(TaskSnapshot).where(TaskSnapshot.id == snapshot_id))
    ).scalar_one_or_none()
    if not snap:
        raise APIError(404, "Snapshot not found", "NOT_FOUND")

    task = (
        await db.execute(select(TaskItem).where(TaskItem.id == snap.task_id))
    ).scalar_one_or_none()
    if not task:
        raise APIError(404, "Task not found", "NOT_FOUND")

    await enforce_shop_access(db, current_user, task.shop_id, "EDIT")

    # Optimistic locking
    if snap.version != body.version:
        return JSONResponse(
            status_code=409,
            content={
                "detail": "Snapshot modified by another user. Reload and retry.",
                "code": "CONFLICT_VERSION",
                "current_version": snap.version,
            },
        )

    before = _snap_to_dict(snap)

    if body.status is not None:
        snap.status = validate_status(body.status)

    # MH decrease check (§7.2.7)
    if body.mh_incurred_hours is not None:
        await check_mh_decrease(
            db, snap, body.mh_incurred_hours, current_user,
            task.shop_id, body.correction_reason,
        )
        snap.mh_incurred_hours = body.mh_incurred_hours

    if body.has_issue is not None:
        snap.has_issue = body.has_issue

    # §7.2.4 deadline_date: explicit null = clear
    if "deadline_date" in body.model_fields_set:
        snap.deadline_date = body.deadline_date

    if body.remarks is not None:
        snap.remarks = body.remarks
    if body.critical_issue is not None:
        snap.critical_issue = body.critical_issue
    if body.correction_reason is not None:
        snap.correction_reason = body.correction_reason

    now = datetime.now(timezone.utc)
    snap.version += 1
    snap.last_updated_at = now
    snap.last_updated_by = current_user["user_id"]
    snap.supervisor_updated_at = now  # §8.4.4

    await db.flush()

    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_snapshot",
        entity_id=snap.id,
        action="UPDATE",
        before=before,
        after=_snap_to_dict(snap),
    )
    await db.commit()

    return {
        "snapshot_id": snap.id,
        "version": snap.version,
        "status": snap.status,
        "mh_incurred_hours": snap.mh_incurred_hours,
        "deadline_date": snap.deadline_date,
        "remarks": snap.remarks,
        "critical_issue": snap.critical_issue,
        "has_issue": snap.has_issue,
        "correction_reason": snap.correction_reason,
        "last_updated_at": snap.last_updated_at,
        "last_updated_by": snap.last_updated_by,
        "supervisor_updated_at": snap.supervisor_updated_at,
    }


# ── 5. PATCH /api/tasks/snapshots/{id}/delete ───────────────────────

@router.patch("/snapshots/{snapshot_id}/delete", response_model=SnapshotDeleteResponse)
async def soft_delete_snapshot(
    snapshot_id: int,
    body: SnapshotVersionRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    snap = (
        await db.execute(select(TaskSnapshot).where(TaskSnapshot.id == snapshot_id))
    ).scalar_one_or_none()
    if not snap:
        raise APIError(404, "Snapshot not found", "NOT_FOUND")

    task = (
        await db.execute(select(TaskItem).where(TaskItem.id == snap.task_id))
    ).scalar_one_or_none()
    if not task:
        raise APIError(404, "Task not found", "NOT_FOUND")

    await enforce_shop_access(db, current_user, task.shop_id, "MANAGE")

    if snap.version != body.version:
        return JSONResponse(
            status_code=409,
            content={
                "detail": "Snapshot modified by another user. Reload and retry.",
                "code": "CONFLICT_VERSION",
                "current_version": snap.version,
            },
        )

    before = _snap_to_dict(snap)
    now = datetime.now(timezone.utc)

    snap.is_deleted = True
    snap.deleted_at = now
    snap.deleted_by = current_user["user_id"]
    snap.version += 1
    snap.last_updated_at = now
    snap.last_updated_by = current_user["user_id"]

    await db.flush()

    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_snapshot",
        entity_id=snap.id,
        action="DELETE",
        before=before,
        after=_snap_to_dict(snap),
    )
    await db.commit()

    return {
        "snapshot_id": snap.id,
        "is_deleted": True,
        "version": snap.version,
        "deleted_at": snap.deleted_at,
        "deleted_by": snap.deleted_by,
    }


# ── 6. PATCH /api/tasks/snapshots/{id}/restore ──────────────────────

@router.patch("/snapshots/{snapshot_id}/restore", response_model=SnapshotRestoreResponse)
async def restore_snapshot(
    snapshot_id: int,
    body: SnapshotVersionRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    snap = (
        await db.execute(select(TaskSnapshot).where(TaskSnapshot.id == snapshot_id))
    ).scalar_one_or_none()
    if not snap:
        raise APIError(404, "Snapshot not found", "NOT_FOUND")

    task = (
        await db.execute(select(TaskItem).where(TaskItem.id == snap.task_id))
    ).scalar_one_or_none()
    if not task:
        raise APIError(404, "Task not found", "NOT_FOUND")

    await enforce_shop_access(db, current_user, task.shop_id, "MANAGE")

    if snap.version != body.version:
        return JSONResponse(
            status_code=409,
            content={
                "detail": "Snapshot modified by another user. Reload and retry.",
                "code": "CONFLICT_VERSION",
                "current_version": snap.version,
            },
        )

    before = _snap_to_dict(snap)
    now = datetime.now(timezone.utc)

    snap.is_deleted = False
    snap.deleted_at = None
    snap.deleted_by = None
    snap.version += 1
    snap.last_updated_at = now
    snap.last_updated_by = current_user["user_id"]

    await db.flush()

    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_snapshot",
        entity_id=snap.id,
        action="RESTORE",
        before=before,
        after=_snap_to_dict(snap),
    )
    await db.commit()

    return {
        "snapshot_id": snap.id,
        "is_deleted": False,
        "version": snap.version,
        "deleted_at": None,
        "deleted_by": None,
    }


# ── 7. POST /api/tasks (create task) ────────────────────────────────

@router.post("", response_model=TaskCreateResponse, status_code=201)
async def create_task(
    body: TaskCreate,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    # Distribution action must stay ADMIN-only.
    if body.assigned_supervisor_id is not None:
        _require_admin(current_user)

    await enforce_shop_access(db, current_user, body.shop_id, "EDIT")

    status = validate_status(body.status)

    # Validate FK: aircraft
    ac = (await db.execute(select(Aircraft).where(Aircraft.id == body.aircraft_id))).scalar_one_or_none()
    if not ac:
        raise APIError(422, "Aircraft not found", "VALIDATION_ERROR", field="aircraft_id")

    # Validate FK: shop
    shop = (await db.execute(select(Shop).where(Shop.id == body.shop_id))).scalar_one_or_none()
    if not shop:
        raise APIError(422, "Shop not found", "VALIDATION_ERROR", field="shop_id")

    # Validate FK: work_package (optional)
    rfo_no = None
    if body.work_package_id is not None:
        wp = (await db.execute(select(WorkPackage).where(WorkPackage.id == body.work_package_id))).scalar_one_or_none()
        if not wp:
            raise APIError(422, "Work package not found", "VALIDATION_ERROR", field="work_package_id")
        rfo_no = wp.rfo_no

    if body.assigned_supervisor_id is not None:
        await _validate_assigned_supervisor(
            db,
            assigned_supervisor_id=body.assigned_supervisor_id,
            shop_id=body.shop_id,
        )

    now = datetime.now(timezone.utc)

    # §8.4.3: assigned_supervisor_id → distributed_at = NOW()
    distributed_at = now if body.assigned_supervisor_id is not None else None

    task = TaskItem(
        aircraft_id=body.aircraft_id,
        shop_id=body.shop_id,
        work_package_id=body.work_package_id,
        assigned_supervisor_id=body.assigned_supervisor_id,
        distributed_at=distributed_at,
        planned_mh=body.planned_mh,
        task_text=body.task_text,
        created_by=current_user["user_id"],
        created_at=now,
    )
    db.add(task)
    await db.flush()

    snap = TaskSnapshot(
        task_id=task.id,
        meeting_date=body.meeting_date,
        status=status,
        mh_incurred_hours=body.mh_incurred_hours,
        remarks=body.remarks or None,
        critical_issue=body.critical_issue or None,
        has_issue=body.has_issue,
        deadline_date=body.deadline_date,
        version=1,
        last_updated_by=current_user["user_id"],
        last_updated_at=now,
        created_at=now,
    )
    db.add(snap)
    await db.flush()

    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_item",
        entity_id=task.id,
        action="CREATE",
        after={
            "task_id": task.id,
            "aircraft_id": task.aircraft_id,
            "shop_id": task.shop_id,
            "work_package_id": task.work_package_id,
            "assigned_supervisor_id": task.assigned_supervisor_id,
            "planned_mh": str(task.planned_mh) if task.planned_mh else None,
            "task_text": task.task_text,
        },
    )
    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_snapshot",
        entity_id=snap.id,
        action="CREATE",
        after=_snap_to_dict(snap),
    )
    await db.commit()

    return {
        "task_id": task.id,
        "snapshot_id": snap.id,
        "meeting_date": snap.meeting_date,
        "shop_id": task.shop_id,
        "aircraft_id": task.aircraft_id,
        "work_package_id": task.work_package_id,
        "rfo_no": rfo_no,
        "assigned_supervisor_id": task.assigned_supervisor_id,
        "assigned_worker_id": task.assigned_worker_id,
        "distributed_at": task.distributed_at,
        "planned_mh": task.planned_mh,
        "task_text": task.task_text,
        "status": snap.status,
        "mh_incurred_hours": snap.mh_incurred_hours,
        "deadline_date": snap.deadline_date,
        "remarks": snap.remarks,
        "critical_issue": snap.critical_issue,
        "has_issue": snap.has_issue,
        "correction_reason": snap.correction_reason,
        "version": snap.version,
        "supervisor_updated_at": snap.supervisor_updated_at,
        "last_updated_at": snap.last_updated_at,
        "last_updated_by": snap.last_updated_by,
    }


# ── 8. POST /api/tasks/import ────────────────────────────────────────

@router.post("/import")
async def import_tasks(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_admin(current_user)

    content = await file.read()
    filename = file.filename or ""

    rows_data: list[dict] = []

    if filename.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, header in enumerate(headers):
                row_dict[header] = row[i] if i < len(row) else None
            rows_data.append(row_dict)
        wb.close()
    else:
        # Default: CSV
        text_content = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text_content))
        for row in reader:
            # Normalize keys to lowercase
            rows_data.append({k.strip().lower(): v for k, v in row.items()})

    # Build lookup maps for validation
    all_ac_regs = {r.strip() for row in rows_data if (r := row.get("ac_reg")) is not None}
    all_rfo_nos = {r.strip() for row in rows_data if (r := row.get("rfo_no")) is not None and r.strip()}

    ac_map: dict[str, Aircraft] = {}
    if all_ac_regs:
        acs = (await db.execute(select(Aircraft).where(Aircraft.ac_reg.in_(all_ac_regs)))).scalars().all()
        ac_map = {a.ac_reg: a for a in acs}

    wp_map: dict[str, WorkPackage] = {}
    if all_rfo_nos:
        wps = (await db.execute(select(WorkPackage).where(WorkPackage.rfo_no.in_(all_rfo_nos)))).scalars().all()
        wp_map = {w.rfo_no: w for w in wps}

    preview: list[dict] = []
    valid_count = 0
    error_count = 0

    for idx, row in enumerate(rows_data, start=1):
        ac_reg = (row.get("ac_reg") or "").strip()
        rfo_no = (row.get("rfo_no") or "").strip() or None
        description = (row.get("description") or "").strip()
        planned_mh_raw = row.get("planned_mh")

        try:
            planned_mh = float(planned_mh_raw) if planned_mh_raw is not None and str(planned_mh_raw).strip() != "" else 0
        except (ValueError, TypeError):
            planned_mh = 0

        entry: dict = {
            "row": idx,
            "ac_reg": ac_reg,
            "rfo_no": rfo_no or "",
            "description": description,
            "planned_mh": planned_mh,
        }

        # Validate aircraft
        if ac_reg not in ac_map:
            entry["valid"] = False
            entry["error"] = "Aircraft not found"
            error_count += 1
        elif rfo_no and rfo_no not in wp_map:
            entry["valid"] = False
            entry["error"] = "Work package not found"
            error_count += 1
        else:
            entry["valid"] = True
            valid_count += 1

        preview.append(entry)

    return {
        "preview": preview,
        "valid_count": valid_count,
        "error_count": error_count,
    }


# ── 9. POST /api/tasks/import/confirm ────────────────────────────────

@router.post("/import/confirm")
async def import_confirm(
    body: ImportConfirmRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_admin(current_user)

    shop = (
        await db.execute(select(Shop).where(Shop.id == body.shop_id))
    ).scalar_one_or_none()
    if not shop:
        raise APIError(422, "Shop not found", "VALIDATION_ERROR", field="shop_id")

    now = datetime.now(timezone.utc)
    created_count = 0

    for item in body.items:
        if item.assigned_supervisor_id is not None:
            await _validate_assigned_supervisor(
                db,
                assigned_supervisor_id=item.assigned_supervisor_id,
                shop_id=body.shop_id,
            )

        # Resolve aircraft
        ac = (
            await db.execute(select(Aircraft).where(Aircraft.ac_reg == item.ac_reg))
        ).scalar_one_or_none()
        if not ac:
            raise APIError(422, f"Aircraft not found: {item.ac_reg}", "VALIDATION_ERROR", field="ac_reg")

        # Resolve work package (optional)
        wp_id = None
        if item.rfo_no:
            wp = (
                await db.execute(select(WorkPackage).where(WorkPackage.rfo_no == item.rfo_no))
            ).scalar_one_or_none()
            if not wp:
                raise APIError(422, f"Work package not found: {item.rfo_no}", "VALIDATION_ERROR", field="rfo_no")
            wp_id = wp.id

        distributed_at = now if item.assigned_supervisor_id is not None else None

        task = TaskItem(
            aircraft_id=ac.id,
            shop_id=body.shop_id,
            work_package_id=wp_id,
            assigned_supervisor_id=item.assigned_supervisor_id,
            distributed_at=distributed_at,
            planned_mh=item.planned_mh,
            task_text=item.description,
            created_by=current_user["user_id"],
            created_at=now,
        )
        db.add(task)
        await db.flush()

        snap = TaskSnapshot(
            task_id=task.id,
            meeting_date=body.meeting_date,
            status="NOT_STARTED",
            mh_incurred_hours=Decimal("0"),
            version=1,
            last_updated_by=current_user["user_id"],
            last_updated_at=now,
            created_at=now,
        )
        db.add(snap)
        await db.flush()

        await write_audit(
            db,
            actor_id=current_user["user_id"],
            entity_type="task_item",
            entity_id=task.id,
            action="IMPORT_CREATE",
            after={
                "task_id": task.id,
                "aircraft_id": task.aircraft_id,
                "shop_id": task.shop_id,
                "work_package_id": task.work_package_id,
                "assigned_supervisor_id": task.assigned_supervisor_id,
                "planned_mh": str(task.planned_mh) if task.planned_mh else None,
                "task_text": task.task_text,
            },
        )

        created_count += 1

    await db.commit()
    return {"created_count": created_count}


# ── 10. POST /api/tasks/bulk-assign ──────────────────────────────────

@router.post("/bulk-assign")
async def bulk_assign(
    body: BulkAssignRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_admin(current_user)

    now = datetime.now(timezone.utc)

    tasks = (
        await db.execute(select(TaskItem).where(TaskItem.id.in_(body.task_ids)))
    ).scalars().all()

    if len(tasks) != len(body.task_ids):
        found_ids = {t.id for t in tasks}
        missing = [tid for tid in body.task_ids if tid not in found_ids]
        raise APIError(404, f"Tasks not found: {missing}", "NOT_FOUND")

    validated_shops: set[int] = set()
    for task in tasks:
        if task.shop_id in validated_shops:
            continue
        await _validate_assigned_supervisor(
            db,
            assigned_supervisor_id=body.assigned_supervisor_id,
            shop_id=task.shop_id,
        )
        validated_shops.add(task.shop_id)

    for task in tasks:
        before = {
            "task_id": task.id,
            "assigned_supervisor_id": task.assigned_supervisor_id,
            "distributed_at": str(task.distributed_at) if task.distributed_at else None,
        }

        task.assigned_supervisor_id = body.assigned_supervisor_id
        task.distributed_at = now

        await db.flush()

        await write_audit(
            db,
            actor_id=current_user["user_id"],
            entity_type="task_item",
            entity_id=task.id,
            action="BULK_ASSIGN",
            before=before,
            after={
                "task_id": task.id,
                "assigned_supervisor_id": task.assigned_supervisor_id,
                "distributed_at": str(task.distributed_at),
            },
        )

    await db.commit()
    return {"assigned_count": len(tasks)}


# ── 11. POST /api/tasks/{task_id}/assign ─────────────────────────────

@router.post("/{task_id}/assign")
async def assign_task(
    task_id: int,
    body: AssignRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_admin(current_user)

    task = (
        await db.execute(select(TaskItem).where(TaskItem.id == task_id))
    ).scalar_one_or_none()
    if not task:
        raise APIError(404, "Task not found", "NOT_FOUND")

    if body.shop_id != task.shop_id:
        raise APIError(
            422,
            "shop_id must match task.shop_id",
            "VALIDATION_ERROR",
            field="shop_id",
        )
    await _validate_assigned_supervisor(
        db,
        assigned_supervisor_id=body.assigned_supervisor_id,
        shop_id=task.shop_id,
    )

    now = datetime.now(timezone.utc)

    before = {
        "task_id": task.id,
        "assigned_supervisor_id": task.assigned_supervisor_id,
        "distributed_at": str(task.distributed_at) if task.distributed_at else None,
    }

    task.assigned_supervisor_id = body.assigned_supervisor_id
    task.distributed_at = now

    await db.flush()

    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_item",
        entity_id=task.id,
        action="ASSIGN",
        before=before,
        after={
            "task_id": task.id,
            "assigned_supervisor_id": task.assigned_supervisor_id,
            "distributed_at": str(task.distributed_at),
        },
    )
    await db.commit()

    return {
        "task_id": task.id,
        "assigned_supervisor_id": task.assigned_supervisor_id,
        "distributed_at": task.distributed_at,
    }


# ── 12. PATCH /api/tasks/{task_id}/assign-worker ────────────────────

@router.patch("/{task_id}/assign-worker")
async def assign_worker(
    task_id: int,
    body: AssignWorkerRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = (
        await db.execute(select(TaskItem).where(TaskItem.id == task_id))
    ).scalar_one_or_none()
    if not task:
        raise APIError(404, "Task not found", "NOT_FOUND")

    await enforce_shop_access(db, current_user, task.shop_id, "EDIT")

    # Validate worker belongs to same shop (check user_shop_access).
    # Null is allowed to clear assignment.
    if body.assigned_worker_id is not None:
        worker_access = (
            await db.execute(
                select(UserShopAccess).where(
                    UserShopAccess.user_id == body.assigned_worker_id,
                    UserShopAccess.shop_id == task.shop_id,
                )
            )
        ).scalar_one_or_none()

        if not worker_access:
            raise APIError(403, "Worker does not have access to this shop", "SHOP_ACCESS_DENIED")

    before = {
        "task_id": task.id,
        "assigned_worker_id": task.assigned_worker_id,
    }

    task.assigned_worker_id = body.assigned_worker_id

    await db.flush()

    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_item",
        entity_id=task.id,
        action="ASSIGN_WORKER",
        before=before,
        after={
            "task_id": task.id,
            "assigned_worker_id": task.assigned_worker_id,
        },
    )
    await db.commit()

    return {
        "task_id": task.id,
        "assigned_worker_id": task.assigned_worker_id,
    }


# ── 13. PATCH /api/tasks/{id}/deactivate ────────────────────────────

@router.patch("/{task_id}/deactivate", response_model=TaskDeactivateResponse)
async def deactivate_task(
    task_id: int,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = (
        await db.execute(select(TaskItem).where(TaskItem.id == task_id))
    ).scalar_one_or_none()
    if not task:
        raise APIError(404, "Task not found", "NOT_FOUND")

    await enforce_shop_access(db, current_user, task.shop_id, "MANAGE")

    now = datetime.now(timezone.utc)

    before = {
        "task_id": task.id,
        "is_active": task.is_active,
        "deactivated_at": str(task.deactivated_at) if task.deactivated_at else None,
        "deactivated_by": task.deactivated_by,
    }

    task.is_active = False
    task.deactivated_at = now
    task.deactivated_by = current_user["user_id"]

    await db.flush()

    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_item",
        entity_id=task.id,
        action="DEACTIVATE",
        before=before,
        after={
            "task_id": task.id,
            "is_active": False,
            "deactivated_at": str(now),
            "deactivated_by": current_user["user_id"],
        },
    )
    await db.commit()

    return {
        "task_id": task.id,
        "is_active": False,
        "deactivated_at": task.deactivated_at,
        "deactivated_by": task.deactivated_by,
    }


# ── 14. PATCH /api/tasks/{id}/reactivate ────────────────────────────

@router.patch("/{task_id}/reactivate", response_model=TaskDeactivateResponse)
async def reactivate_task(
    task_id: int,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = (
        await db.execute(select(TaskItem).where(TaskItem.id == task_id))
    ).scalar_one_or_none()
    if not task:
        raise APIError(404, "Task not found", "NOT_FOUND")

    await enforce_shop_access(db, current_user, task.shop_id, "MANAGE")

    before = {
        "task_id": task.id,
        "is_active": task.is_active,
        "deactivated_at": str(task.deactivated_at) if task.deactivated_at else None,
        "deactivated_by": task.deactivated_by,
    }

    task.is_active = True
    task.deactivated_at = None
    task.deactivated_by = None

    await db.flush()

    await write_audit(
        db,
        actor_id=current_user["user_id"],
        entity_type="task_item",
        entity_id=task.id,
        action="REACTIVATE",
        before=before,
        after={
            "task_id": task.id,
            "is_active": True,
            "deactivated_at": None,
            "deactivated_by": None,
        },
    )
    await db.commit()

    return {
        "task_id": task.id,
        "is_active": True,
        "deactivated_at": None,
        "deactivated_by": None,
    }


# ── 15. POST /api/tasks/init-week ───────────────────────────────────

@router.post("/init-week", response_model=InitWeekResponse)
async def init_week_endpoint(
    body: InitWeekRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await enforce_shop_access(db, current_user, body.shop_id, "MANAGE")

    result = await init_week(db, body.shop_id, body.meeting_date, current_user["user_id"])

    if result["created_count"] > 0:
        await write_audit(
            db,
            actor_id=current_user["user_id"],
            entity_type="task_snapshot",
            entity_id=0,
            action="INIT_WEEK",
            after=result,
        )

    await db.commit()
    return result
